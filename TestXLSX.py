import openpyxl
import xml.etree.ElementTree as ET
import random
import DB_Hardcoded as db_value


#Loading exel file
wb = openpyxl.load_workbook('newXML.xlsx', 'r')
sheet1 = wb['Sheet1']
row = sheet1.max_row
column = sheet1.max_column

rate_card_datapoint = []
iterator_of_ratecard_datapoint = 0

for i in range(2, row + 1):
    if(str(sheet1.cell(i, 1).value) != 'None'):
        rate_card_datapoint.append(i)
        #rate_card_datapoint_including_last_row.append(i)

rate_card_datapoint.append(row + 1)



print('Number of datapoint', rate_card_datapoint)

#Testing
print(row)
print(column)
print(sheet1.cell(2, 14).value)
print(sheet1.cell(3, 14).value)
print(sheet1.cell(4, 14).value)


for i in range(2, row + 1):
    print(str(sheet1.cell(i, 15).value))

print(rate_card_datapoint)

#XML Elements declaration
price_list = ET.Element('price_list')
price_list.attrib = {'xmlns:xsi': 'http://www.w3.org/2001/XMLSchema-instance', 'version':'7.4', 'xsi:noNamespaceSchemaLocation':'../../setup/scripts/price_list.xsd'}
# plan = ET.Element('plan')
# plan.attrib = {'ondemand_billing': 'no'}
# deal = ET.Element('deal')
# deal.attrib = {'customization_flag': 'optional', 'ondemand_billing': 'no', 'grant_resources_as_group': 'no'}
# product = ET.Element('product')
# product.attrib = {'type' : 'subscription', 'partial' : 'no'}

rate_card_old = 'rate_card'
service_deal_old = ET.Element('test')

deal_array_elem_appender = [1]

flag = 1

#Generating XML Document
for i in rate_card_datapoint:
    print(i)


    if(i < row):

        if sheet1.cell(i, 1).value == rate_card_old:

            RANDOM_RANDINT = random.randint(100, 999)

            deal_array_elem = ET.Element('deal_array_elem')
            deal_array_elem.attrib = {'type': 'optional'}
            deal_array_elem.text = str(sheet1.cell(i, 1).value) + " " + str(sheet1.cell(i, 4).value) + " " + str(sheet1.cell(i, 5).value)

            if flag == 1:
                for item in deal_array_elem_appender:
                    service_deal_old.append(item)
                    flag = 0

            service_deal_old.append(deal_array_elem)

            deal = ET.Element('deal')
            deal.attrib = {'customization_flag': 'optional', 'ondemand_billing': 'no', 'grant_resources_as_group': 'no'}

            #Populating deal child element

            deal_name = ET.Element('deal_name')
            deal_name.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(
                sheet1.cell(i, 5).value)

            deal_code = ET.Element('deal_code')
            deal_code.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(
                sheet1.cell(i, 5).value)

            description_deal = ET.Element('description')
            description_deal.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(
                sheet1.cell(i, 5).value)

            start_time = ET.Element('start_time')

            datetime = ET.Element('datetime')

            date = ET.Element('date')
            date.attrib = {'yyyy': '1970', 'mm': '01', 'dd': '01'}

            time = ET.Element('time')
            time.attrib = {'hh': '00', 'mm': '00', 'ss': '00', 'tzhr': '0', 'tzmn': '00'}

            datetime.append(date)
            datetime.append(time)

            start_time.append(datetime)

            permitted = ET.Element('permitted')
            permitted.text = str(sheet1.cell(i, 8).value)

            deal_product = ET.Element('deal_product')
            deal_product.attrib = {'status': 'active'}

            product_name = ET.Element('product_name')
            product_name.text = str(sheet1.cell(i, 1).value) + " " + str(sheet1.cell(i, 4).value) + " " + str(
                sheet1.cell(i, 5).value)

            product_code = ET.Element('product_code')
            product_code.text = str(sheet1.cell(i, 1).value) + " " + str(sheet1.cell(i, 4).value) + " " + str(
                sheet1.cell(i, 5).value)

            quantity = ET.Element('quantity')
            quantity.text = '1.0'

            purchase_discount = ET.Element('purchase_discount')
            purchase_discount.text = '0.0'

            cycle_discount = ET.Element('cycle_discount')
            cycle_discount.text = '0.0'

            usage_discount = ET.Element('usage_discount')
            usage_discount.text = '0.0'

            status_flag = ET.Element('status_flag')
            status_flag.text = '0'

            deal_product.append(product_name)
            deal_product.append(product_code)
            deal_product.append(quantity)
            deal_product.append(purchase_discount)
            deal_product.append(cycle_discount)
            deal_product.append(usage_discount)
            deal_product.append(status_flag)

            deal.append(deal_name)
            deal.append(deal_code)
            deal.append(description)
            deal.append(start_time)
            deal.append(permitted)
            deal.append(deal_product)

            price_list.append(deal)

            product = ET.Element('product')
            product.attrib = {'type': 'subscription', 'partial': 'no'}

            #Populating child elements of product

            product_name_product = ET.Element('product_name')
            product_name_product.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(
                sheet1.cell(i, 5).value)

            product_code_product = ET.Element('product_code')
            product_code_product.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(
                sheet1.cell(i, 5).value)

            priority_product = ET.Element('priority')
            priority_product.text = '0'

            description_product = ET.Element('description')
            description_product.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(
                sheet1.cell(i, 5).value)

            permitted_product = ET.Element('permitted')
            permitted_product.text = str(sheet1.cell(i, 8).value)

            event_rating_map = ET.Element('event_rating_map')
            event_rating_map.attrib = {'tod_mode': 'start_time', 'timezone_mode': 'event', 'min_unit': 'none',
                                       'incr_unit': 'none', 'rounding_rule': 'nearest', 'flag': 'not_rate_cancelled'}

            event_type_product = ET.Element('event_type')
            event_type_product.text = str(sheet1.cell(i, 9).value)

            rum_name = ET.Element('rum_name')
            rum_name.text = str(sheet1.cell(i, 10).value)

            min_quantity = ET.Element('min_quantity')
            min_quantity.text = '0'

            incr_quantity = ET.Element('incr_quantity')
            incr_quantity.text = '1'

            event_rating_map.append(event_type_product)
            event_rating_map.append(rum_name)
            event_rating_map.append(min_quantity)
            event_rating_map.append(incr_quantity)

            rate_plan_selector = ET.Element('rate_plan_selector')

            # Populating rate plan selector child attributes
            rate_plan_selector_name = ET.Element('rate_plan_selector_name')
            rate_plan_selector_name.text = str(sheet1.cell(i, 1).value) + ' ' + str(RANDOM_RANDINT)

            rate_plan_selector_code = ET.Element('rate_plan_selector_code')
            rate_plan_selector_code.text = str(sheet1.cell(i, 1).value) + ' ' + str(RANDOM_RANDINT)

            selector_rate_plan_selector = ET.Element('selector')

            column_selector = ET.Element('column')
            column_selector.attrib = {'operator': 'equal', 'separator': ';'}

            field_name_column_selector = ET.Element('field_name')
            s1 = db_value.evenName_atributeCode_mapping.get(str(sheet1.cell(i, 9).value))
            field_name_column_selector.text = db_value.newDict.get(s1)

            column_selector.append(field_name_column_selector)

           # print(rate_card_datapoint)
            attribute_values_datapoint = []
            for j in range(i, rate_card_datapoint[iterator_of_ratecard_datapoint + 1]):
                if (str(sheet1.cell(j, 14).value) != 'None'):
                    attribute_values_datapoint.append(j)

            print(attribute_values_datapoint)

            # Populating selector
            selector_rate_plan_selector.append(column_selector)

            # dynamic product->event_rating_map->selector
            for j in attribute_values_datapoint:
                value_range_selector = ET.Element('value_range')
                value_selector = ET.Element('value')
                value_selector.text = str(sheet1.cell(j, 14).value)

                rate_plan_name_selector = ET.Element('rate_plan_name')
                rate_plan_name_selector.text = str(sheet1.cell(i, 1).value) + ' ' + str(
                    sheet1.cell(j, 14).value) + ' ' + str(RANDOM_RANDINT)
                rate_plan_code_selector = ET.Element('rate_plan_code')
                rate_plan_code_selector.text = str(sheet1.cell(i, 1).value) + ' ' + str(
                    sheet1.cell(j, 14).value) + ' ' + str(RANDOM_RANDINT)
                impact_category_selector = ET.Element('impact_category')
                impact_category_selector.text = 'default'

                value_range_selector.append(value_selector)
                value_range_selector.append(rate_plan_name_selector)
                value_range_selector.append(rate_plan_code_selector)
                value_range_selector.append(impact_category_selector)

                selector_rate_plan_selector.append(value_range_selector)

            # Populating rate plan selector
            rate_plan_selector.append(rate_plan_selector_name)
            rate_plan_selector.append(rate_plan_selector_code)
            rate_plan_selector.append(selector_rate_plan_selector)
            event_rating_map.append(rate_plan_selector)

            for j in attribute_values_datapoint:
                rate_plan_event_rating_map = ET.Element('rate_plan')
                rate_plan_event_rating_map.attrib = {'tax_when': 'now'}

                curr_id = ET.Element('currency_id')
                curr_id.text = db_value.currencyName_currencyCode_mapping.get(str(sheet1.cell(i, 5).value))

                rate_plan_name_rate_plan = ET.Element('rate_plan_name')
                rate_plan_name_rate_plan.text = str(sheet1.cell(i, 1).value) + ' ' + str(
                    sheet1.cell(j, 14).value) + ' ' + str(
                    RANDOM_RANDINT)

                rate_plan_code_rate_plan = ET.Element('rate_plan_code')
                rate_plan_code_rate_plan.text = str(sheet1.cell(i, 1).value) + ' ' + str(
                    sheet1.cell(j, 14).value) + ' ' + str(
                    RANDOM_RANDINT)

                event_type_rate_plan = ET.Element('event_type')
                event_type_rate_plan.text = str(sheet1.cell(i, 9).value)

                tax_code_rate_plan = ET.Element('tax_code')
                tax_code_rate_plan.text = str(sheet1.cell(i, 11).value)

                advance_billing_offset_rate_plan = ET.Element('advance_billing_offset')
                advance_billing_offset_rate_plan.text = '0'

                cycle_fee_flags_rate_plan = ET.Element('cycle_fee_flags')
                cycle_fee_flags_rate_plan.text = '0'

                rate_plan_event_rating_map.append(rate_plan_name_rate_plan)
                rate_plan_event_rating_map.append(rate_plan_code_rate_plan)
                rate_plan_event_rating_map.append(curr_id)
                rate_plan_event_rating_map.append(event_type_rate_plan)
                rate_plan_event_rating_map.append(tax_code_rate_plan)
                rate_plan_event_rating_map.append(advance_billing_offset_rate_plan)
                rate_plan_event_rating_map.append(cycle_fee_flags_rate_plan)

                rate_tier_rate_plan = ET.Element('rate_tier')
                rate_tier_rate_plan.attrib = {'date_range_type': 'absolute'}

                rate_tier_name_rate_tier = ET.Element('rate_tier_name')
                rate_tier_name_rate_tier.text = 'RateTier_0'

                priority_rate_tier = ET.Element('priority')
                priority_rate_tier.text = '100'

                date_range_rate_tier = ET.Element('date_range')
                date_range_rate_tier.attrib = {'relative_start_unit': 'none', 'relative_end_unit': 'none'}

                date_range_name_date_range = ET.Element('date_range_name')
                date_range_name_date_range.text = 'DateRange_1'

                relative_start_offset_date_range = ET.Element('relative_start_offset')
                relative_start_offset_date_range.text = '0'

                relative_end_offset_date_range = ET.Element('relative_end_offset')
                relative_end_offset_date_range.text = '0'

                rate_date_range = ET.Element('rate')
                rate_date_range.attrib = {'type': 'default', 'step_type': 'total_quantity_rated',
                                          'prorate_first': 'prorate', 'prorate_last': 'prorate'}

                description_rate = ET.Element('description')
                description_rate.text = 'Pricing'

                rate_date_range.append(description_rate)

                step_resource_id = ET.Element('step_resource_id')
                step_resource_id.text = '0'

                rate_date_range.append(step_resource_id)

                k = j

                while str(sheet1.cell(k, 17).value) != '999999':
                    quantity_tier_rate = ET.Element('quantity_tier')
                    step_min = ET.Element('step_min')
                    step_min.text = str(sheet1.cell(k, 16).value)

                    step_max = ET.Element('step_max')
                    step_max.text = str(sheet1.cell(k, 17).value)

                    balance_impact_quantity_tier = ET.Element('balance_impact')
                    balance_impact_quantity_tier.attrib = {'scaled_unit': 'none', 'flag': 'discountable proratable'}

                    resource_id_balance_impact = ET.Element('resource_id')
                    resource_id_balance_impact.text = db_value.currencyName_currencyCode_mapping.get(
                        str(sheet1.cell(i, 5).value))

                    impact_category_balance_impact = ET.Element('impact_category')
                    impact_category_balance_impact.text = 'default'

                    glid = ET.Element('glid')
                    glid.text = db_value.tier_glid_mapping.get(str(sheet1.cell(k, 15).value))

                    fixed_amount_balance_impact = ET.Element('fixed_amount')
                    fixed_amount_balance_impact.text = '0'

                    scaled_amount = ET.Element('scaled_amount')
                    scaled_amount.text = str(sheet1.cell(k, 18).value)

                    balance_impact_quantity_tier.append(resource_id_balance_impact)
                    balance_impact_quantity_tier.append(impact_category_balance_impact)
                    balance_impact_quantity_tier.append(glid)
                    balance_impact_quantity_tier.append(fixed_amount_balance_impact)
                    balance_impact_quantity_tier.append(scaled_amount)

                    quantity_tier_rate.append(step_min)
                    quantity_tier_rate.append(step_max)
                    quantity_tier_rate.append(balance_impact_quantity_tier)

                    # rate_date_range.append(description_rate)
                    # rate_date_range.append(step_resource_id)
                    rate_date_range.append(quantity_tier_rate)

                    k += 1

                if (str(sheet1.cell(k, 17).value) == '999999'):
                    quantity_tier_rate_2 = ET.Element('quantity_tier')
                    step_min = ET.Element('step_min')
                    step_min.text = str(sheet1.cell(k, 16).value)

                    balance_impact_quantity_tier_2 = ET.Element('balance_impact')
                    balance_impact_quantity_tier_2.attrib = {'scaled_unit': 'none', 'flag': 'discountable proratable'}

                    resource_id_2 = ET.Element('resource_id')
                    resource_id_2.text = db_value.currencyName_currencyCode_mapping.get(str(sheet1.cell(i, 5).value))

                    balance_impact_quantity_tier_2.append(resource_id_2)

                    impact_category_2 = ET.Element('impact_category')
                    impact_category_2.text = 'default'

                    glid_2 = ET.Element('glid')
                    glid_2.text = db_value.tier_glid_mapping.get(str(sheet1.cell(k, 15).value))

                    fixed_amount_2 = ET.Element('fixed_amount')
                    fixed_amount_2.text = '0'

                    scaled_amount_2 = ET.Element('scaled_amount')
                    scaled_amount_2.text = str(sheet1.cell(k, 18).value)

                    # balance_impact_quantity_tier_2.append(resource_id_2)
                    balance_impact_quantity_tier_2.append(impact_category_2)
                    balance_impact_quantity_tier_2.append(glid_2)
                    balance_impact_quantity_tier_2.append(fixed_amount_2)
                    balance_impact_quantity_tier_2.append(scaled_amount_2)

                    quantity_tier_rate_2.append(step_min)
                    quantity_tier_rate_2.append(balance_impact_quantity_tier_2)

                    rate_date_range.append(quantity_tier_rate_2)

                date_range_rate_tier.append(date_range_name_date_range)
                date_range_rate_tier.append(relative_start_offset_date_range)
                date_range_rate_tier.append(relative_end_offset_date_range)
                date_range_rate_tier.append(rate_date_range)
                rate_tier_rate_plan.append(rate_tier_name_rate_tier)
                rate_tier_rate_plan.append(priority_rate_tier)
                rate_tier_rate_plan.append(date_range_rate_tier)

                rate_plan_event_rating_map.append(rate_tier_rate_plan)
                event_rating_map.append(rate_plan_event_rating_map)

            if (iterator_of_ratecard_datapoint < len(rate_card_datapoint) - 1):
                iterator_of_ratecard_datapoint += 1

            product.append(product_name_product)
            product.append(product_code_product)
            product.append(priority_product)
            product.append(description_product)
            product.append(permitted_product)
            product.append(event_rating_map)

            price_list.append(product)

            continue

        RANDOM_RANDINT = random.randint(100, 999)

        rate_card_old = sheet1.cell(i, 1).value
        print('OUT Continue')
       # print(rate_card_old)
        print(service_deal_old.text)
        plan = ET.Element('plan')
        plan.attrib = {'ondemand_billing': 'no'}
        deal = ET.Element('deal')
        deal.attrib = {'customization_flag': 'optional', 'ondemand_billing': 'no', 'grant_resources_as_group': 'no'}
        product = ET.Element('product')
        product.attrib = {'type': 'subscription', 'partial': 'no'}

        #Populating plan child elements
        plan_name = ET.Element('plan_name')
        plan_name.text = str(sheet1.cell(i, 1).value)

        plan_code = ET.Element('plan_code')
        plan_code.text = str(sheet1.cell(i, 1).value)

        description = ET.Element('description')
        description.text = str(sheet1.cell(i, 1).value)

        service_deal = ET.Element('service_deal')

        service_name = ET.Element('service_name')
        service_name.text = str(sheet1.cell(i, 8).value)

        service_id = ET.Element('service_id')
        service_id.text = str(sheet1.cell(i, 8).value) + " " + str(RANDOM_RANDINT)

        deal_array_elem = ET.Element('deal_array_elem')
        deal_array_elem.attrib = {'type': 'optional'}
        deal_array_elem.text = str(sheet1.cell(i, 1).value) + " " + str(sheet1.cell(i, 4).value) + " " + str(sheet1.cell(i, 5).value)

        # bal_info_index = ET.Element('bal_info_index')
        # bal_info_index.text = '0'

        # subscription_index = ET.Element('subscription_index')
        # subscription_index.text = '0'

        service_deal.append(service_name)
        service_deal.append(service_id)
        if str(sheet1.cell(i, 1).value) == str(sheet1.cell(rate_card_datapoint[iterator_of_ratecard_datapoint + 1], 1).value):
            deal_array_elem_appender.clear()
            deal_array_elem_appender.append(deal_array_elem)
            flag = 1
        else:
            service_deal.append(deal_array_elem)

        # service_deal.append(bal_info_index)
        # service_deal.append(subscription_index)
        service_deal_old = service_deal

        price_list.append(plan)

        #Populating deal child elements
        deal_name = ET.Element('deal_name')
        deal_name.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(sheet1.cell(i, 5).value)

        deal_code = ET.Element('deal_code')
        deal_code.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(sheet1.cell(i, 5).value)

        description_deal = ET.Element('description')
        description_deal.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(sheet1.cell(i, 5).value)

        start_time = ET.Element('start_time')

        datetime = ET.Element('datetime')

        date = ET.Element('date')
        date.attrib = {'yyyy': '1970', 'mm':'01', 'dd': '01'}

        time = ET.Element('time')
        time.attrib = {'hh': '00', 'mm':'00', 'ss': '00', 'tzhr': '0', 'tzmn': '00'}

        datetime.append(date)
        datetime.append(time)

        start_time.append(datetime)

        permitted = ET.Element('permitted')
        permitted.text = str(sheet1.cell(i, 8).value)

        deal_product = ET.Element('deal_product')
        deal_product.attrib = {'status':'active'}

        product_name = ET.Element('product_name')
        product_name.text = str(sheet1.cell(i, 1).value) + " " + str(sheet1.cell(i, 4).value) + " " + str(sheet1.cell(i, 5).value)

        product_code = ET.Element('product_code')
        product_code.text = str(sheet1.cell(i, 1).value) + " " + str(sheet1.cell(i, 4).value) + " " + str(sheet1.cell(i, 5).value)

        quantity = ET.Element('quantity')
        quantity.text = '1.0'

        purchase_discount = ET.Element('purchase_discount')
        purchase_discount.text = '0.0'

        cycle_discount = ET.Element('cycle_discount')
        cycle_discount.text = '0.0'

        usage_discount = ET.Element('usage_discount')
        usage_discount.text = '0.0'

        status_flag = ET.Element('status_flag')
        status_flag.text = '0'

        deal_product.append(product_name)
        deal_product.append(product_code)
        deal_product.append(quantity)
        deal_product.append(purchase_discount)
        deal_product.append(cycle_discount)
        deal_product.append(usage_discount)
        deal_product.append(status_flag)

        #Populating product child elements
        product_name_product = ET.Element('product_name')
        product_name_product.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(sheet1.cell(i, 5).value)

        product_code_product = ET.Element('product_code')
        product_code_product.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(sheet1.cell(i, 5).value)

        priority_product = ET.Element('priority')
        priority_product.text = '0'

        description_product = ET.Element('description')
        description_product.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(i, 4).value) + ' ' + str(sheet1.cell(i, 5).value)

        permitted_product = ET.Element('permitted')
        permitted_product.text = str(sheet1.cell(i, 8).value)

        event_rating_map = ET.Element('event_rating_map')
        event_rating_map.attrib = {'tod_mode' : 'start_time', 'timezone_mode' : 'event', 'min_unit':'none', 'incr_unit':'none', 'rounding_rule':'nearest', 'flag':'not_rate_cancelled'}

        event_type_product = ET.Element('event_type')
        event_type_product.text = str(sheet1.cell(i, 9).value)

        rum_name = ET.Element('rum_name')
        rum_name.text = str(sheet1.cell(i, 10).value)

        min_quantity = ET.Element('min_quantity')
        min_quantity.text = '0'

        incr_quantity = ET.Element('incr_quantity')
        incr_quantity.text = '1'

        event_rating_map.append(event_type_product)
        event_rating_map.append(rum_name)
        event_rating_map.append(min_quantity)
        event_rating_map.append(incr_quantity)

        rate_plan_selector = ET.Element('rate_plan_selector')

        #Populating rate plan selector child attributes
        rate_plan_selector_name = ET.Element('rate_plan_selector_name')
        rate_plan_selector_name.text = str(sheet1.cell(i, 1).value) + ' ' + str(RANDOM_RANDINT)

        rate_plan_selector_code = ET.Element('rate_plan_selector_code')
        rate_plan_selector_code.text = str(sheet1.cell(i, 1).value) + ' ' + str(RANDOM_RANDINT)

        selector_rate_plan_selector = ET.Element('selector')

        column_selector = ET.Element('column')
        column_selector.attrib = {'operator':'equal', 'separator':';'}

        field_name_column_selector = ET.Element('field_name')
        s1 = db_value.evenName_atributeCode_mapping.get(str(sheet1.cell(i, 9).value))
        field_name_column_selector.text = db_value.newDict.get(s1)

        column_selector.append(field_name_column_selector)

       # print(rate_card_datapoint)
        attribute_values_datapoint = []
        for j in range(i, rate_card_datapoint[iterator_of_ratecard_datapoint + 1]):
            if(str(sheet1.cell(j, 14).value) != 'None'):
                attribute_values_datapoint.append(j)

        #print(attribute_values_datapoint)

        # Populating selector
        selector_rate_plan_selector.append(column_selector)

        #dynamic product->event_rating_map->selector
        for j in attribute_values_datapoint:

            value_range_selector = ET.Element('value_range')
            value_selector = ET.Element('value')
            value_selector.text = str(sheet1.cell(j, 14).value)

            rate_plan_name_selector = ET.Element('rate_plan_name')
            rate_plan_name_selector.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(j, 14).value) + ' ' + str(RANDOM_RANDINT)
            rate_plan_code_selector = ET.Element('rate_plan_code')
            rate_plan_code_selector.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(j, 14).value) + ' ' + str(RANDOM_RANDINT)
            impact_category_selector = ET.Element('impact_category')
            impact_category_selector.text = 'default'

            value_range_selector.append(value_selector)
            value_range_selector.append(rate_plan_name_selector)
            value_range_selector.append(rate_plan_code_selector)
            value_range_selector.append(impact_category_selector)

            selector_rate_plan_selector.append(value_range_selector)

        # Populating rate plan selector
        rate_plan_selector.append(rate_plan_selector_name)
        rate_plan_selector.append(rate_plan_selector_code)
        rate_plan_selector.append(selector_rate_plan_selector)
        event_rating_map.append(rate_plan_selector)

        for j in attribute_values_datapoint:
            rate_plan_event_rating_map = ET.Element('rate_plan')
            rate_plan_event_rating_map.attrib = {'tax_when':'now'}

            curr_id = ET.Element('currency_id')
            curr_id.text = db_value.currencyName_currencyCode_mapping.get(str(sheet1.cell(i, 5).value))


            rate_plan_name_rate_plan = ET.Element('rate_plan_name')
            rate_plan_name_rate_plan.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(j, 14).value) + ' ' + str(
                RANDOM_RANDINT)

            rate_plan_code_rate_plan = ET.Element('rate_plan_code')
            rate_plan_code_rate_plan.text = str(sheet1.cell(i, 1).value) + ' ' + str(sheet1.cell(j, 14).value) + ' ' + str(
                RANDOM_RANDINT)

            event_type_rate_plan = ET.Element('event_type')
            event_type_rate_plan.text = str(sheet1.cell(i, 9).value)

            tax_code_rate_plan = ET.Element('tax_code')
            tax_code_rate_plan.text = str(sheet1.cell(i, 11).value)

            advance_billing_offset_rate_plan = ET.Element('advance_billing_offset')
            advance_billing_offset_rate_plan.text = '0'

            cycle_fee_flags_rate_plan = ET.Element('cycle_fee_flags')
            cycle_fee_flags_rate_plan.text = '0'

            rate_plan_event_rating_map.append(rate_plan_name_rate_plan)
            rate_plan_event_rating_map.append(rate_plan_code_rate_plan)
            rate_plan_event_rating_map.append(curr_id)
            rate_plan_event_rating_map.append(event_type_rate_plan)
            rate_plan_event_rating_map.append(tax_code_rate_plan)
            rate_plan_event_rating_map.append(advance_billing_offset_rate_plan)
            rate_plan_event_rating_map.append(cycle_fee_flags_rate_plan)

            rate_tier_rate_plan = ET.Element('rate_tier')
            rate_tier_rate_plan.attrib = {'date_range_type' : 'absolute'}

            rate_tier_name_rate_tier = ET.Element('rate_tier_name')
            rate_tier_name_rate_tier.text = 'RateTier_0'

            priority_rate_tier = ET.Element('priority')
            priority_rate_tier.text = '100'

            date_range_rate_tier = ET.Element('date_range')
            date_range_rate_tier.attrib = {'relative_start_unit': 'none', 'relative_end_unit': 'none'}

            date_range_name_date_range = ET.Element('date_range_name')
            date_range_name_date_range.text = 'DateRange_1'

            relative_start_offset_date_range = ET.Element('relative_start_offset')
            relative_start_offset_date_range.text = '0'

            relative_end_offset_date_range = ET.Element('relative_end_offset')
            relative_end_offset_date_range.text = '0'

            rate_date_range = ET.Element('rate')
            rate_date_range.attrib = {'type': 'default', 'step_type': 'total_quantity_rated', 'prorate_first': 'prorate', 'prorate_last': 'prorate'}

            description_rate = ET.Element('description')
            description_rate.text = 'Pricing'

            rate_date_range.append(description_rate)

            step_resource_id = ET.Element('step_resource_id')
            step_resource_id.text = '0'

            rate_date_range.append(step_resource_id)

            k = j

            while str(sheet1.cell(k, 17).value) != '999999':
                quantity_tier_rate = ET.Element('quantity_tier')
                step_min = ET.Element('step_min')
                step_min.text = str(sheet1.cell(k, 16).value)

                step_max = ET.Element('step_max')
                step_max.text = str(sheet1.cell(k, 17).value)

                balance_impact_quantity_tier = ET.Element('balance_impact')
                balance_impact_quantity_tier.attrib = {'scaled_unit': 'none', 'flag': 'discountable proratable'}

                resource_id_balance_impact = ET.Element('resource_id')
                resource_id_balance_impact.text = db_value.currencyName_currencyCode_mapping.get(str(sheet1.cell(i, 5).value))

                impact_category_balance_impact = ET.Element('impact_category')
                impact_category_balance_impact.text = 'default'

                glid = ET.Element('glid')
                glid.text = db_value.tier_glid_mapping.get(str(sheet1.cell(k, 15).value))

                fixed_amount_balance_impact = ET.Element('fixed_amount')
                fixed_amount_balance_impact.text = '0'

                scaled_amount = ET.Element('scaled_amount')
                scaled_amount.text = str(sheet1.cell(k, 18).value)

                balance_impact_quantity_tier.append(resource_id_balance_impact)
                balance_impact_quantity_tier.append(impact_category_balance_impact)
                balance_impact_quantity_tier.append(glid)
                balance_impact_quantity_tier.append(fixed_amount_balance_impact)
                balance_impact_quantity_tier.append(scaled_amount)

                quantity_tier_rate.append(step_min)
                quantity_tier_rate.append(step_max)
                quantity_tier_rate.append(balance_impact_quantity_tier)

                # rate_date_range.append(description_rate)
                # rate_date_range.append(step_resource_id)
                rate_date_range.append(quantity_tier_rate)

                k += 1

            if(str(sheet1.cell(k, 17).value) == '999999'):

                quantity_tier_rate_2 = ET.Element('quantity_tier')
                step_min = ET.Element('step_min')
                step_min.text = str(sheet1.cell(k, 16).value)

                balance_impact_quantity_tier_2 = ET.Element('balance_impact')
                balance_impact_quantity_tier_2.attrib = {'scaled_unit': 'none', 'flag': 'discountable proratable'}

                resource_id_2 = ET.Element('resource_id')
                resource_id_2.text = db_value.currencyName_currencyCode_mapping.get(str(sheet1.cell(i, 5).value))

                balance_impact_quantity_tier_2.append(resource_id_2)

                impact_category_2 = ET.Element('impact_category')
                impact_category_2.text = 'default'

                glid_2 = ET.Element('glid')
                glid_2.text = db_value.tier_glid_mapping.get(str(sheet1.cell(k, 15).value))

                fixed_amount_2 = ET.Element('fixed_amount')
                fixed_amount_2.text = '0'

                scaled_amount_2 = ET.Element('scaled_amount')
                scaled_amount_2.text = str(sheet1.cell(k, 18).value)

                #balance_impact_quantity_tier_2.append(resource_id_2)
                balance_impact_quantity_tier_2.append(impact_category_2)
                balance_impact_quantity_tier_2.append(glid_2)
                balance_impact_quantity_tier_2.append(fixed_amount_2)
                balance_impact_quantity_tier_2.append(scaled_amount_2)

                quantity_tier_rate_2.append(step_min)
                quantity_tier_rate_2.append(balance_impact_quantity_tier_2)

                rate_date_range.append(quantity_tier_rate_2)

            date_range_rate_tier.append(date_range_name_date_range)
            date_range_rate_tier.append(relative_start_offset_date_range)
            date_range_rate_tier.append(relative_end_offset_date_range)
            date_range_rate_tier.append(rate_date_range)
            rate_tier_rate_plan.append(rate_tier_name_rate_tier)
            rate_tier_rate_plan.append(priority_rate_tier)
            rate_tier_rate_plan.append(date_range_rate_tier)

            rate_plan_event_rating_map.append(rate_tier_rate_plan)
            event_rating_map.append(rate_plan_event_rating_map)

        if(iterator_of_ratecard_datapoint < len(rate_card_datapoint) - 1):
            iterator_of_ratecard_datapoint += 1

        # event_rating_map.append(event_type_product)
        # event_rating_map.append(rum_name)
        # event_rating_map.append(min_quantity)
        # event_rating_map.append(incr_quantity)

        #Populating deal
        deal.append(deal_name)
        deal.append(deal_code)
        deal.append(description_deal)
        deal.append(start_time)
        deal.append(permitted)
        deal.append(deal_product)

        #Populating plan
        plan.append(plan_name)
        plan.append(plan_code)
        plan.append(description)
        plan.append(service_deal_old)

        #Populating product
        product.append(product_name_product)
        product.append(product_code_product)
        product.append(priority_product)
        product.append(description_product)
        product.append(permitted_product)
        product.append(event_rating_map)

       # price_list.append(plan)
        price_list.append(deal)
        price_list.append(product)

tree = ET.ElementTree(price_list)
with open('Customer4.xml', 'wb') as files:
    tree.write(files)
